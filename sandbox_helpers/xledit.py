"""Excel fayllarni FORMATNI BUZMASDAN tahrirlash.

Nega kerak edi: GPT yozgan oddiy kod .xls ni xlutils.copy bilan nusxalab
katakchaga yozganda ikki narsa yo'qoladi —
  1) katakchaning XF indeksi (shrift, rang, chegara, tekislash, sana
     formati) — natijada qiymat "General, Arial 10, pastki-chap" bo'lib
     qoladi;
  2) kitobning maxsus rang palitrasi — pushti sarlavha sariq bo'lib
     ketadi.
Bu modul ikkalasini ham tiklaydi. .xlsx uchun esa umuman qayta yozilmaydi:
ZIP ichidagi bitta XML tugun almashtiriladi, qolgan hamma narsa (rasm,
diagramma, shartli formatlash, pivot) bayt-bayt saqlanadi.

Ishlatish:
    import xledit
    for c in xledit.cells('input.xls'):      # tuzilishni ko'rish
        print(c)
    xledit.replace('input.xls', 'output/natija.xls', '31.12.99', 0)
    # yoki aniq manzil bo'yicha:
    xledit.edit('input.xls', 'output/natija.xls', {'D5': 0})
"""
from __future__ import annotations

import os
import re
import shutil
import zipfile

__all__ = ["cells", "edit", "replace"]

_ADDR_RE = re.compile(r"^(?:(.+)!)?([A-Za-z]{1,3})(\d+)$")


# ---------------------------------------------------------------- manzil

def a1_to_rc(addr: str) -> tuple[int, int]:
    """'D5' -> (4, 3). 0 dan boshlanadigan (qator, ustun)."""
    m = _ADDR_RE.match(addr.strip())
    if not m:
        raise ValueError(f"noto'g'ri katakcha manzili: {addr!r}")
    col = 0
    for ch in m.group(2).upper():
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(3)) - 1, col - 1


def rc_to_a1(row: int, col: int) -> str:
    """(4, 3) -> 'D5'."""
    name, col = "", col + 1
    while col:
        col, rem = divmod(col - 1, 26)
        name = chr(65 + rem) + name
    return f"{name}{row + 1}"


def _norm_changes(changes: dict, sheet_names: list[str]) -> dict:
    """Har xil kalitlarni {(sheet_idx, row, col): value} ga keltiradi.

    Qabul qilinadi: 'D5', 'Sheet1!D5', (row, col), (sheet_idx, row, col).
    """
    out = {}
    for key, val in changes.items():
        if isinstance(key, str):
            m = _ADDR_RE.match(key.strip())
            if not m:
                raise ValueError(f"noto'g'ri manzil: {key!r}")
            sheet = 0
            if m.group(1):
                want = m.group(1).strip("'\"")
                if want not in sheet_names:
                    raise ValueError(f"varaq topilmadi: {want!r}")
                sheet = sheet_names.index(want)
            row, col = a1_to_rc(m.group(2) + m.group(3))
        elif len(key) == 2:
            sheet, (row, col) = 0, key
        elif len(key) == 3:
            sheet, row, col = key
        else:
            raise ValueError(f"noto'g'ri kalit: {key!r}")
        out[(sheet, row, col)] = val
    return out


# ------------------------------------------------------------------ .xls

def _xls_book(path: str):
    import xlrd

    try:
        return xlrd.open_workbook(path, formatting_info=True), True
    except NotImplementedError:
        # ba'zi fayllarda formatting_info qo'llab-quvvatlanmaydi
        return xlrd.open_workbook(path), False


def _xls_cells(path: str) -> list[dict]:
    import xlrd

    book, _ = _xls_book(path)
    out = []
    for si in range(book.nsheets):
        sh = book.sheet_by_index(si)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                item = {
                    "sheet": si, "sheet_name": sh.name,
                    "row": r, "col": c, "addr": rc_to_a1(r, c),
                    "value": cell.value, "is_date": cell.ctype == xlrd.XL_CELL_DATE,
                }
                if item["is_date"]:
                    try:
                        item["date"] = xlrd.xldate.xldate_as_datetime(
                            cell.value, book.datemode)
                    except Exception:
                        pass
                out.append(item)
    return out


def _xls_copy(book):
    """xlutils.copy o'rniga — nusxa BILAN BIRGA uslublar ro'yxatini qaytaradi.

    Oddiy xlutils.copy() faqat kitobni beradi, uslublarni emas; keyin
    ws.write() katakchani standart uslub bilan yozib yuboradi. XF
    indeksini qo'lda qaytarib ham bo'lmaydi — nusxalashda indekslar
    siljiydi. XLWTWriter esa asl XF indeksidan xlwt uslubiga xarita beradi.
    """
    from xlutils.filter import XLRDReader, XLWTWriter, process

    writer = XLWTWriter()
    process(XLRDReader(book, "input.xls"), writer)
    return writer.output[0][1], writer.style_list


def _xls_edit(src: str, dst: str, changes: dict) -> int:
    book, has_fmt = _xls_book(src)
    names = [book.sheet_by_index(i).name for i in range(book.nsheets)]
    todo = _norm_changes(changes, names)

    out, styles = _xls_copy(book)

    # maxsus rang palitrasini tiklash — xlutils buni ko'chirmaydi
    if has_fmt:
        for idx, rgb in (book.colour_map or {}).items():
            if rgb is not None and 8 <= idx <= 63:
                try:
                    out.set_colour_RGB(idx, *rgb)
                except Exception:
                    pass

    done = 0
    for (si, r, c), value in todo.items():
        style = None
        if has_fmt:
            xf = book.sheet_by_index(si).cell_xf_index(r, c)
            if xf < len(styles):
                style = styles[xf]
        ws = out.get_sheet(si)
        ws.write(r, c, value, style) if style else ws.write(r, c, value)
        done += 1

    os.makedirs(os.path.dirname(dst) or ".", exist_ok=True)
    out.save(dst)
    return done


# ----------------------------------------------------------------- .xlsx

def _xlsx_sheet_parts(zf: zipfile.ZipFile) -> list[tuple[str, str]]:
    """[(varaq nomi, zip ichidagi yo'l)] — workbook tartibida."""
    book = zf.read("xl/workbook.xml").decode("utf-8", "replace")
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
    # atribut tartibi turlicha bo'ladi (Excel Id'ni, openpyxl Target'ni
    # oldin yozadi), shuning uchun har bir tegni alohida o'qiymiz
    by_id = {}
    for tag in re.findall(r"<Relationship\b[^>]*>", rels):
        rid = re.search(r'\bId="([^"]+)"', tag)
        target = re.search(r'\bTarget="([^"]+)"', tag)
        if rid and target:
            by_id[rid.group(1)] = target.group(1)
    out = []
    for tag in re.findall(r"<sheet\b[^>]*/?>", book):
        name = re.search(r'name="([^"]*)"', tag)
        rid = re.search(r'r:id="([^"]+)"', tag)
        if not (name and rid and rid.group(1) in by_id):
            continue
        target = by_id[rid.group(1)].lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        out.append((_unescape(name.group(1)), target))
    return out


def _unescape(s: str) -> str:
    return (s.replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'")
             .replace("&amp;", "&"))


def _escape(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;"))


def _xlsx_cells(path: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = []
    try:
        for si, ws in enumerate(wb.worksheets):
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    out.append({
                        "sheet": si, "sheet_name": ws.title,
                        "row": cell.row - 1, "col": cell.column - 1,
                        "addr": cell.coordinate, "value": cell.value,
                        "is_date": hasattr(cell.value, "year"),
                    })
    finally:
        wb.close()
    return out


def _new_cell_xml(addr: str, style_attr: str, value) -> str:
    if isinstance(value, bool):
        return f'<c r="{addr}"{style_attr} t="b"><v>{int(value)}</v></c>'
    if isinstance(value, (int, float)):
        return f'<c r="{addr}"{style_attr}><v>{value}</v></c>'
    if value is None or value == "":
        return f'<c r="{addr}"{style_attr}/>'
    txt = _escape(str(value))
    return (f'<c r="{addr}"{style_attr} t="inlineStr">'
            f'<is><t xml:space="preserve">{txt}</t></is></c>')


def _patch_sheet_xml(xml: str, cell_changes: dict) -> tuple[str, int]:
    """cell_changes: {'D5': value}. Faqat shu tugunlar almashtiriladi."""
    done = 0
    for addr, value in cell_changes.items():
        pat = re.compile(
            r'<c\b[^>]*\br="%s"(?:\s[^>]*)?(?:/>|>.*?</c>)' % re.escape(addr),
            re.S)
        m = pat.search(xml)
        if not m:
            continue  # bo'sh katakcha — XML'da umuman yo'q
        style = re.search(r'\ss="\d+"', m.group(0))
        new = _new_cell_xml(addr, style.group(0) if style else "", value)
        xml = xml[:m.start()] + new + xml[m.end():]
        done += 1
    return xml, done


def _xlsx_edit(src: str, dst: str, changes: dict) -> int:
    zin = zipfile.ZipFile(src)
    try:
        parts = _xlsx_sheet_parts(zin)
        todo = _norm_changes(changes, [n for n, _ in parts])

        by_part: dict[str, dict] = {}
        for (si, r, c), value in todo.items():
            if si >= len(parts):
                raise ValueError(f"varaq indeksi mavjud emas: {si}")
            by_part.setdefault(parts[si][1], {})[rc_to_a1(r, c)] = value

        patched, done = {}, 0
        for part, cell_changes in by_part.items():
            xml = zin.read(part).decode("utf-8")
            xml, n = _patch_sheet_xml(xml, cell_changes)
            patched[part], done = xml.encode("utf-8"), done + n

        os.makedirs(os.path.dirname(dst) or ".", exist_ok=True)
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = patched.get(info.filename)
                if data is None:
                    data = zin.read(info.filename)
                zout.writestr(info, data)
        return done
    finally:
        zin.close()


# ------------------------------------------------------------ ommaviy API

def cells(path: str) -> list[dict]:
    """Fayldagi bo'sh bo'lmagan katakchalar ro'yxati.

    Har biri: sheet, sheet_name, row, col, addr ('D5'), value, is_date.
    Tahrirlashdan oldin shu bilan kerakli katakcha manzilini toping.
    """
    if path.lower().endswith(".xls"):
        return _xls_cells(path)
    return _xlsx_cells(path)


def edit(src: str, dst: str, changes: dict) -> int:
    """Ko'rsatilgan katakchalarni almashtirib, boshqa HAMMA NARSANI saqlaydi.

    changes kalitlari: 'D5', 'Varaq1!D5', (qator, ustun) yoki
    (varaq_indeksi, qator, ustun). Qator/ustun 0 dan boshlanadi.
    Nechta katakcha yozilganini qaytaradi.
    """
    if not changes:
        shutil.copy2(src, dst)
        return 0
    if src.lower().endswith(".xls"):
        return _xls_edit(src, dst, changes)
    return _xlsx_edit(src, dst, changes)


def replace(src: str, dst: str, old, new, *, sheet=None) -> int:
    """`old` qiymatiga teng har bir katakchani `new` ga almashtiradi.

    Taqqoslash matn ko'rinishida bo'ladi, ya'ni replace(..., '31.12.99', 0)
    ham matn, ham sana katakchasini topadi. Almashtirilgan soni qaytadi.
    """
    want = str(old).strip()
    changes = {}
    for c in cells(src):
        if sheet is not None and c["sheet"] != sheet:
            continue
        got = c["value"]
        texts = {str(got).strip()}
        if c.get("date"):
            d = c["date"]
            texts |= {d.strftime("%d.%m.%y"), d.strftime("%d.%m.%Y"),
                      d.strftime("%d/%m/%Y"), d.strftime("%Y-%m-%d")}
        elif c["is_date"] and hasattr(got, "strftime"):
            texts |= {got.strftime("%d.%m.%y"), got.strftime("%d.%m.%Y"),
                      got.strftime("%d/%m/%Y"), got.strftime("%Y-%m-%d")}
        if isinstance(got, float) and got.is_integer():
            texts.add(str(int(got)))
        if want in texts:
            changes[(c["sheet"], c["row"], c["col"])] = new
    return edit(src, dst, changes)


# ------------------------------------------------------------- self-check

if __name__ == "__main__":
    import tempfile

    import xlrd
    import xlwt

    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "src.xls")

    wb = xlwt.Workbook()
    xlwt.add_palette_colour("pushti", 0x22)
    wb.set_colour_RGB(0x22, 0xFC, 0xE4, 0xD6)
    ws = wb.add_sheet("Ma'lumot")
    st = xlwt.easyxf(
        "font: bold on, height 280; align: horiz center, vert center;"
        "pattern: pattern solid, fore_colour pushti;"
        "borders: left thin, right thin, top thin, bottom thin")
    ws.write_merge(0, 0, 0, 3, "SARLAVHA", st)
    ws.row(2).height_mismatch, ws.row(2).height = True, 600
    ws.col(1).width = 6000
    ws.write(2, 0, "Aliyev Ali", st)
    ws.write(2, 1, "31.12.99", st)
    ws.write(3, 0, "Valiyev Vali", st)
    ws.write(3, 1, "15.03.01", st)
    wb.save(src)

    assert a1_to_rc("D5") == (4, 3) and rc_to_a1(4, 3) == "D5"
    assert a1_to_rc("AA1") == (0, 26) and rc_to_a1(0, 26) == "AA1"

    got = {c["addr"]: c["value"] for c in cells(src)}
    assert got["B3"] == "31.12.99", got

    dst = os.path.join(tmp, "out.xls")
    assert replace(src, dst, "31.12.99", 0) == 1

    a = xlrd.open_workbook(src, formatting_info=True)
    b = xlrd.open_workbook(dst, formatting_info=True)
    sa, sb = a.sheet_by_index(0), b.sheet_by_index(0)

    assert sb.cell(2, 1).value == 0, sb.cell(2, 1).value
    assert sb.cell(3, 1).value == "15.03.01", "boshqa katakcha o'zgardi"
    assert sb.cell(2, 0).value == "Aliyev Ali"

    def look(bk, sh, r, c):
        """Katakcha ko'rinishi — XF indeksi emas, haqiqiy formatlash.
        (nusxalashda indekslar siljiydi, shuning uchun indeksni emas,
        u ko'rsatadigan uslubni taqqoslaymiz)"""
        xf = bk.xf_list[sh.cell_xf_index(r, c)]
        f = bk.font_list[xf.font_index]
        return (bk.format_map[xf.format_key].format_str, f.bold, f.height,
                f.name, xf.alignment.hor_align, xf.alignment.vert_align,
                xf.background.background_colour_index,
                xf.background.fill_pattern, xf.border.left_line_style,
                xf.border.top_line_style)

    # asosiy tekshiruv: HAR BIR katakcha ko'rinishi o'zgarmadimi
    for r in range(sa.nrows):
        for c in range(sa.ncols):
            assert look(a, sa, r, c) == look(b, sb, r, c), \
                f"({r},{c}) uslubi o'zgardi"
    assert b.colour_map[0x22] == a.colour_map[0x22], "palitra yo'qoldi"
    assert sb.merged_cells == sa.merged_cells, "birlashtirish yo'qoldi"
    assert sb.rowinfo_map[2].height == sa.rowinfo_map[2].height
    assert sb.colinfo_map[1].width == sa.colinfo_map[1].width

    # .xlsx: ZIP tugun almashtirish
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    xsrc = os.path.join(tmp, "src.xlsx")
    w2 = openpyxl.Workbook()
    s2 = w2.active
    s2["A1"] = "Ism"
    s2["B1"] = "Sana"
    s2["B1"].font = Font(bold=True, size=14, color="FF0000")
    s2["B1"].fill = PatternFill("solid", fgColor="FCE4D6")
    s2["A2"], s2["B2"] = "Ali", "31.12.99"
    s2["B2"].font = Font(bold=True, size=14)
    s2["B2"].fill = PatternFill("solid", fgColor="FCE4D6")
    s2.merge_cells("A4:C4")
    s2.column_dimensions["B"].width = 30
    w2.save(xsrc)

    xdst = os.path.join(tmp, "out.xlsx")
    assert replace(xsrc, xdst, "31.12.99", 0) == 1
    r2 = openpyxl.load_workbook(xdst)
    o2 = r2.active
    assert o2["B2"].value == 0, o2["B2"].value
    assert o2["B2"].font.bold and o2["B2"].font.sz == 14, "xlsx uslubi yo'qoldi"
    assert o2["B2"].fill.fgColor.rgb.endswith("FCE4D6"), "xlsx to'ldirish yo'qoldi"
    assert "A4:C4" in {str(r) for r in o2.merged_cells.ranges}
    assert o2.column_dimensions["B"].width == 30
    # o'zgarmagan qismlar bayt-bayt bir xilmi
    za, zb = zipfile.ZipFile(xsrc), zipfile.ZipFile(xdst)
    same = [n for n in za.namelist() if "worksheets" not in n]
    assert all(za.read(n) == zb.read(n) for n in same), "boshqa qismlar o'zgardi"
    za.close(); zb.close()

    # matnga almashtirish
    xdst2 = os.path.join(tmp, "out2.xlsx")
    assert edit(xsrc, xdst2, {"B2": "yangi & <matn>"}) == 1
    assert openpyxl.load_workbook(xdst2).active["B2"].value == "yangi & <matn>"

    shutil.rmtree(tmp, ignore_errors=True)
    print("xledit: barcha tekshiruvlar o'tdi.")
